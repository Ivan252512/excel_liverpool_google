from openpyxl import load_workbook

class Excel():
    
    def __init__(self, src):
        self.excel = load_workbook(src, data_only=True)
        self.name = str(src)

    def get_pages(self):
        return self.excel.sheetnames

    def get_all_columns(self):
        all_columns = []
        for j in self.get_pages():
            sheet_column = [j, []]
            sheet = self.excel[j]
            for cell in sheet[1]:
                sheet_column[1].append(cell.value)
            all_columns.append(sheet_column)
        return all_columns            

    def load_as_array(self):
        all_columns = []
        for j in self.get_pages():
            sheet_column = [j, []]
            sheet = self.excel[j]
            sheet_column[1].append(["vacio"])
            for i in range(1,sheet.max_row):
                sheet_column[1].append([cell.value for cell in sheet[i]])
            all_columns.append(sheet_column)
        return all_columns   

"""
if __name__ == "__main__":
    excel = Excel("ArchivosP.xlsx")

    all_columns = excel.load_as_array()

    documents_to_save = {
        "name" : "Archivos"
    }

    pages_to_save = {
        "document" : "Archivos",
        "name": "",
    }

    row_to_save = {
        "page" : "",
    }

    data_to_save = {
        "page" : "",
    }

    sig=False
    title = []
    for i in all_columns:
        pages_to_save["name"] = i[0]
        for j in i[1]:
            row_to_save["page"] = i[0]
            if sig:
                title = j
                sig=False
            if j[0] == "vacio":
                sig=True
            else:
                data_to_save["page"] = i[0]
                for x in range(len(title)):
                    data_to_save[title[x]] = j[x]

            print(data_to_save)
"""